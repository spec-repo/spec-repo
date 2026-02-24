#!/usr/bin/env python3
"""
요구사항정의서 Excel 스타일 공통 모듈.
create-template.py, json-to-excel.py 에서 import한다.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 스타일 상수 ────────────────────────────────────────────────
HEADER_FILL  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
HEADER_FONT  = Font(bold=True, size=9)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
WRAP_ALIGN   = Alignment(wrap_text=True, vertical="top")
THIN         = Side(style="thin")
THIN_BORDER  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def style_header(cell):
    cell.fill      = HEADER_FILL
    cell.font      = HEADER_FONT
    cell.alignment = CENTER_ALIGN
    cell.border    = THIN_BORDER


def style_data(cell):
    cell.alignment = WRAP_ALIGN
    cell.border    = THIN_BORDER
    cell.font      = Font(size=9)


def build_sheet(wb, sheet_name, col_defs, merges_r1, sample_row=None):
    """
    시트 생성 및 헤더 작성.

    col_defs:  [(col, r1_label, r2_label_or_None, width), ...]
    merges_r1: {r1_label: (start_col, end_col), ...}
    sample_row: 선택적 샘플 데이터 행 (list). None 이면 샘플 행 없음.
    """
    ws = wb.create_sheet(sheet_name)

    for col, _, _, width in col_defs:
        ws.column_dimensions[get_column_letter(col)].width = width

    seen_r1 = set()
    for col, r1, r2, _ in col_defs:
        c1 = ws.cell(row=1, column=col)
        c2 = ws.cell(row=2, column=col)
        if r1 not in seen_r1:
            c1.value = r1
            seen_r1.add(r1)
        if r2:
            c2.value = r2
        style_header(c1)
        style_header(c2)

    for label, (start_col, end_col) in merges_r1.items():
        ws.merge_cells(
            start_row=1, start_column=start_col,
            end_row=1, end_column=end_col,
        )
        cell = ws.cell(row=1, column=start_col)
        cell.value = label
        style_header(cell)

    single_cols = [col for col, _, r2, _ in col_defs if r2 is None]
    for col in single_cols:
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 28

    if sample_row is not None:
        for c_idx, val in enumerate(sample_row, start=1):
            cell = ws.cell(row=3, column=c_idx, value=val)
            style_data(cell)
        ws.row_dimensions[3].height = 40

    ws.freeze_panes = "A3"
    return ws
