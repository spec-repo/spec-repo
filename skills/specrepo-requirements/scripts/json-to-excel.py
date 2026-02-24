#!/usr/bin/env python3
"""
requirements.json → xlsx 재생성 스크립트.

Usage:
    uv run --with openpyxl python3 json-to-excel.py
    uv run --with openpyxl python3 json-to-excel.py --input 01-requirements/requirements.json
    uv run --with openpyxl python3 json-to-excel.py --out snapshots/requirements/export.xlsx
    uv run --with openpyxl python3 json-to-excel.py --version 1.2.3
"""

import sys
import json
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    print("Run: uv run --with openpyxl python3 json-to-excel.py", file=sys.stderr)
    sys.exit(1)

sys.path.insert(0, str(Path(__file__).parent))
from sheet_styles import build_sheet, style_data, THIN_BORDER


# ── 컬럼 정의 (create-template.py 와 동일) ────────────────────
FUNC_COLS = [
    ( 1, "요구사항 ID",            None,                            10),
    ( 2, "제안요청서",              "요구사항 명",                    16),
    ( 3, "제안요청서",              "요구사항 설명",                  20),
    ( 4, "세부요구사항ID",          None,                            12),
    ( 5, "업무기능",               "대분류",                         10),
    ( 6, "업무기능",               "중분류",                         10),
    ( 7, "고객요구사항",            None,                            20),
    ( 8, "상세요구사항",            None,                            28),
    ( 9, "설계 및 구현방안",        "전제조건 및 구현방안",            22),
    (10, "설계 및 구현방안",        "제약사항",                       16),
    (11, "산출물",                 None,                             8),
    (12, "기능구분",               None,                             8),
    (13, "우선순위",               None,                             8),
    (14, "출처",                   None,                            10),
    (15, "요구자",                 None,                            10),
    (16, "요청일",                 None,                            12),
    (17, "완료일",                 None,                            12),
    (18, "업무담당자",             None,                            10),
    (19, "동료검토",               None,                            10),
    (20, "담당자연락처",            None,                            14),
    (21, "적용구분",               None,                            10),
    (22, "요구관리상태",            None,                            12),
    (23, "변경근거",               None,                            14),
]
FUNC_MERGES_R1 = {
    "제안요청서":      (2, 3),
    "업무기능":        (5, 6),
    "설계 및 구현방안": (9, 10),
}

NONFUNC_COLS = [
    ( 1, "요구사항 ID",            None,                            10),
    ( 2, "제안요청서",              "요구사항 명",                    16),
    ( 3, "제안요청서",              "요구사항 설명",                  20),
    ( 4, "요구사항ID(세부)",        None,                            12),
    ( 5, "업무기능",               "대분류",                         10),
    ( 6, "업무기능",               "중분류",                         10),
    ( 7, "고객요구사항",            None,                            20),
    ( 8, "상세요구사항",            None,                            28),
    ( 9, "상세구현방안",            None,                            22),
    (10, "산출물명",               None,                             8),
    (11, "관련근거",               None,                            10),
    (12, "기능구분",               None,                             8),
    (13, "우선순위",               None,                             8),
    (14, "출처",                   None,                            10),
    (15, "요구자",                 None,                            10),
    (16, "요청일",                 None,                            12),
    (17, "완료일",                 None,                            12),
    (18, "업무담당자",             None,                            10),
    (19, "동료검토",               None,                            10),
    (20, "담당자연락처",            None,                            14),
    (21, "적용구분",               None,                            10),
    (22, "요구관리상태",            None,                            12),
    (23, "변경근거",               None,                            14),
    (24, "설계 및 구현 제약사항",   "전제조건",                       18),
    (25, "설계 및 구현 제약사항",   "제약사항",                       16),
]
NONFUNC_MERGES_R1 = {
    "제안요청서":           (2, 3),
    "업무기능":             (5, 6),
    "설계 및 구현 제약사항": (24, 25),
}

# JSON 필드 → 컬럼 번호 (기능)
FUNC_PARENT_COLS = {"id": 1, "rfp_name": 2, "rfp_desc": 3}
FUNC_SUB_COLS = {
    "sub_id": 4, "category_major": 5, "category_minor": 6,
    "customer_requirement": 7, "detail_requirement": 8,
    "implementation": 9, "constraints": 10, "deliverable": 11,
    "func_type": 12, "priority": 13, "source": 14,
    "requester": 15, "request_date": 16, "completion_date": 17,
    "owner": 18, "peer_review": 19, "contact": 20,
    "apply_type": 21, "status": 22, "change_reason": 23,
}

# JSON 필드 → 컬럼 번호 (비기능)
NONFUNC_PARENT_COLS = {"id": 1, "rfp_name": 2, "rfp_desc": 3}
NONFUNC_SUB_COLS = {
    "sub_id": 4, "category_major": 5, "category_minor": 6,
    "customer_requirement": 7, "detail_requirement": 8,
    "implementation": 9, "deliverable": 10, "related_basis": 11,
    "func_type": 12, "priority": 13, "source": 14,
    "requester": 15, "request_date": 16, "completion_date": 17,
    "owner": 18, "peer_review": 19, "contact": 20,
    "apply_type": 21, "status": 22, "change_reason": 23,
    "design_precondition": 24, "design_constraints": 25,
}

_MERGE_ALIGN = Alignment(wrap_text=True, vertical="top")
_DATA_FONT   = Font(size=9)


def write_data_rows(ws, items, parent_col_map, sub_col_map):
    """
    JSON 데이터를 시트 행으로 기록한다.
    부모 필드(id, rfp_name, rfp_desc)는 세부항목 수만큼 세로 병합 처리.
    기록된 데이터 행 수를 반환한다.
    """
    row = 3
    for item in items:
        subs = item.get("sub_requirements") or [{}]
        n = len(subs)
        start_row = row

        for sub in subs:
            # 부모 필드
            for field, col in parent_col_map.items():
                cell = ws.cell(row=row, column=col, value=item.get(field, ""))
                style_data(cell)
            # 세부 필드
            for field, col in sub_col_map.items():
                cell = ws.cell(row=row, column=col, value=sub.get(field, ""))
                style_data(cell)
            ws.row_dimensions[row].height = 40
            row += 1

        # 부모 필드 세로 병합 (세부항목 2개 이상)
        if n > 1:
            for col in parent_col_map.values():
                ws.merge_cells(
                    start_row=start_row, start_column=col,
                    end_row=row - 1,     end_column=col,
                )
                cell = ws.cell(row=start_row, column=col)
                cell.alignment = _MERGE_ALIGN
                cell.border    = THIN_BORDER
                cell.font      = _DATA_FONT

    return row - 3


def find_json_path(start_dir):
    """01-requirements/requirements.json 탐색."""
    current = Path(start_dir).resolve()
    for _ in range(10):
        candidate = current / "01-requirements" / "requirements.json"
        if candidate.exists():
            return candidate
        if (current / "AGENTS.md").exists() or (current / ".git").is_dir():
            return current / "01-requirements" / "requirements.json"
        parent = current.parent
        if parent == current:
            break
        current = parent
    return Path(start_dir).resolve() / "01-requirements" / "requirements.json"


def main():
    args = sys.argv[1:]

    if "-h" in args or "--help" in args:
        print(__doc__)
        sys.exit(0)

    in_path         = None
    out_path        = None
    version_override = None

    i = 0
    while i < len(args):
        if args[i] == "--input" and i + 1 < len(args):
            in_path = Path(args[i + 1])
            i += 2
        elif args[i] == "--out" and i + 1 < len(args):
            out_path = Path(args[i + 1])
            i += 2
        elif args[i] == "--version" and i + 1 < len(args):
            version_override = args[i + 1]
            i += 2
        else:
            i += 1

    if in_path is None:
        in_path = find_json_path(Path.cwd())

    if not in_path.exists():
        print(f"ERROR: JSON 파일 없음: {in_path}", file=sys.stderr)
        print("먼저 excel-to-json.py 로 임포트하거나 --input 으로 경로를 지정하세요.")
        sys.exit(1)

    with open(in_path, encoding="utf-8") as f:
        data = json.load(f)

    version = version_override or data.get("_meta", {}).get("version") or "latest"

    if out_path is None:
        snapshot_dir = in_path.parent.parent / "snapshots" / "requirements"
        snapshot_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d")
        out_path = snapshot_dir / f"requirements_v{version}_{ts}.xlsx"

    out_path.parent.mkdir(parents=True, exist_ok=True)

    functional    = data.get("functional", [])
    nonfunctional = data.get("nonfunctional", [])

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    ws_func    = build_sheet(wb, "기능_요구사항",   FUNC_COLS,    FUNC_MERGES_R1)
    n_func     = write_data_rows(ws_func,    functional,    FUNC_PARENT_COLS,    FUNC_SUB_COLS)

    ws_nonfunc = build_sheet(wb, "비기능_요구사항", NONFUNC_COLS, NONFUNC_MERGES_R1)
    n_nonfunc  = write_data_rows(ws_nonfunc, nonfunctional, NONFUNC_PARENT_COLS, NONFUNC_SUB_COLS)

    wb.save(out_path)
    print(f"✅ 내보내기 완료: {out_path}")
    print(f"   기능 요구사항:   {len(functional)}개 ({n_func}행)")
    print(f"   비기능 요구사항: {len(nonfunctional)}개 ({n_nonfunc}행)")


if __name__ == "__main__":
    main()
