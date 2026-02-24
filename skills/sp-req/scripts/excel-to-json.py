#!/usr/bin/env python3
"""
요구사항정의서 xlsx → 01-requirements/requirements.json 변환 스크립트.

Usage:
    uv run --with openpyxl python3 excel-to-json.py <excel-file>
    uv run --with openpyxl python3 excel-to-json.py <excel-file> --merge upsert
    uv run --with openpyxl python3 excel-to-json.py <excel-file> --merge replace
    uv run --with openpyxl python3 excel-to-json.py <excel-file> --merge append
    uv run --with openpyxl python3 excel-to-json.py <excel-file> --out path/to/output.json

merge 전략:
    upsert  (기본) ID 기준 덮어쓰기, 기존에만 있는 ID는 유지
    replace         시트 전체 교체
    append          신규 ID만 추가, 기존 ID는 유지
"""

import sys
import json
import re
from datetime import datetime, timezone
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    print("Run: uv run --with openpyxl python3 excel-to-json.py <file>", file=sys.stderr)
    sys.exit(1)


# ── 컬럼 매핑: Excel 헤더 → JSON 필드명 ──────────────────────
# (헤더는 parse-excel.py 방식의 "r1 > r2" 형식을 따름)

# 기능 요구사항 (기능_요구사항 시트)
FUNC_PARENT_FIELDS = {
    "요구사항 ID":               "id",
    "제안요청서 > 요구사항 명":   "rfp_name",
    "제안요청서 > 요구사항 설명": "rfp_desc",
}
FUNC_SUB_FIELDS = {
    "세부요구사항ID":                          "sub_id",
    "업무기능 > 대분류":                       "category_major",
    "업무기능 > 중분류":                       "category_minor",
    "고객요구사항":                            "customer_requirement",
    "상세요구사항":                            "detail_requirement",
    "설계 및 구현방안 > 전제조건 및 구현방안":  "implementation",
    "설계 및 구현방안 > 제약사항":              "constraints",
    "산출물":                                 "deliverable",
    "기능구분":                               "func_type",
    "우선순위":                               "priority",
    "출처":                                   "source",
    "요구자":                                 "requester",
    "요청일":                                 "request_date",
    "완료일":                                 "completion_date",
    "업무담당자":                              "owner",
    "동료검토":                               "peer_review",
    "담당자연락처":                            "contact",
    "적용구분":                               "apply_type",
    "요구관리상태":                            "status",
    "변경근거":                               "change_reason",
}

# 비기능 요구사항 (비기능_요구사항 시트)
NONFUNC_PARENT_FIELDS = {
    "요구사항 ID":               "id",
    "제안요청서 > 요구사항 명":   "rfp_name",
    "제안요청서 > 요구사항 설명": "rfp_desc",
}
NONFUNC_SUB_FIELDS = {
    "요구사항ID(세부)":                         "sub_id",
    "업무기능 > 대분류":                        "category_major",
    "업무기능 > 중분류":                        "category_minor",
    "고객요구사항":                             "customer_requirement",
    "상세요구사항":                             "detail_requirement",
    "상세구현방안":                             "implementation",
    "산출물명":                                "deliverable",
    "관련근거":                                "related_basis",
    "기능구분":                                "func_type",
    "우선순위":                                "priority",
    "출처":                                    "source",
    "요구자":                                  "requester",
    "요청일":                                  "request_date",
    "완료일":                                  "completion_date",
    "업무담당자":                               "owner",
    "동료검토":                                "peer_review",
    "담당자연락처":                             "contact",
    "적용구분":                                "apply_type",
    "요구관리상태":                             "status",
    "변경근거":                                "change_reason",
    "설계 및 구현 제약사항 > 전제조건":          "design_precondition",
    "설계 및 구현 제약사항 > 제약사항":          "design_constraints",
}


# ── 저수준 셀 유틸 ─────────────────────────────────────────────

def resolve_merged_cells(ws):
    """병합된 셀의 값을 모든 병합 범위에 채운다."""
    merge_values = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
        value = top_left.value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_values[(row, col)] = value
    return merge_values


def get_cell_value(ws, row, col, merge_map):
    coord = (row, col)
    if coord in merge_map:
        return merge_map[coord]
    return ws.cell(row=row, column=col).value


def detect_header_rows(ws, merge_map, max_col):
    """1~2행 헤더 파싱. (컬럼명 목록, 데이터 시작 행) 반환."""
    row1 = [get_cell_value(ws, 1, c, merge_map) for c in range(1, max_col + 1)]
    row2 = [get_cell_value(ws, 2, c, merge_map) for c in range(1, max_col + 1)]

    has_subheader = any(
        row2[i] and row2[i] != row1[i] for i in range(len(row1))
    )

    if has_subheader:
        headers = []
        for i in range(max_col):
            r1 = str(row1[i]).strip() if row1[i] else ""
            r2 = str(row2[i]).strip() if row2[i] else ""
            if r2 and r2 != r1:
                headers.append(f"{r1} > {r2}" if r1 else r2)
            else:
                headers.append(r1)
        return headers, 3
    else:
        headers = [str(v).strip() if v else f"col_{i+1}" for i, v in enumerate(row1)]
        return headers, 2


# ── 파싱 로직 ──────────────────────────────────────────────────

def parse_version_from_filename(filepath):
    """파일명에서 버전 추출 (_0.7.8, _v0.7.8, _0.7.8v 패턴). 없으면 None."""
    name = Path(filepath).stem
    m = re.search(r'_v?(\d+\.\d+\.\d+)v?', name)
    return m.group(1) if m else None


def find_project_root(start_dir):
    """AGENTS.md 또는 .git 이 있는 디렉토리를 프로젝트 루트로 반환."""
    current = Path(start_dir).resolve()
    for _ in range(10):
        if (current / "AGENTS.md").exists() or (current / ".git").is_dir():
            return current
        parent = current.parent
        if parent == current:
            break
        current = parent
    return Path(start_dir).resolve()


def parse_requirements_sheet(ws, parent_fields, sub_fields):
    """
    요구사항 시트를 파싱해 normalized 구조로 반환한다.

    Returns: [{"id": ..., "rfp_name": ..., "rfp_desc": ..., "sub_requirements": [...]}, ...]
    """
    max_row, max_col = ws.max_row, ws.max_column
    if max_row < 3 or max_col == 0:
        return []

    merge_map = resolve_merged_cells(ws)
    headers, data_start = detect_header_rows(ws, merge_map, max_col)

    parents = {}
    parent_order = []

    for r in range(data_start, max_row + 1):
        row_vals = [get_cell_value(ws, r, c, merge_map) for c in range(1, max_col + 1)]

        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        row = {}
        for i, h in enumerate(headers):
            if i < len(row_vals):
                val = row_vals[i]
                row[h] = str(val).strip() if val is not None else ""

        parent_id = row.get("요구사항 ID", "").strip()
        if not parent_id:
            continue

        if parent_id not in parents:
            entry = {"id": parent_id, "sub_requirements": []}
            for excel_h, json_f in parent_fields.items():
                if json_f != "id":
                    entry[json_f] = row.get(excel_h, "")
            parents[parent_id] = entry
            parent_order.append(parent_id)

        sub = {}
        for excel_h, json_f in sub_fields.items():
            sub[json_f] = row.get(excel_h, "")
        parents[parent_id]["sub_requirements"].append(sub)

    return [parents[pid] for pid in parent_order]


def detect_sheet_type(sheet_name):
    """시트명으로 functional/nonfunctional 판별."""
    if "비기능" in sheet_name:
        return "nonfunctional"
    if "기능" in sheet_name:
        return "functional"
    return None


def merge_requirements(existing, incoming, strategy):
    """merge 전략 적용."""
    if strategy == "replace":
        return incoming

    existing_map = {item["id"]: i for i, item in enumerate(existing)}

    if strategy == "append":
        result = list(existing)
        for item in incoming:
            if item["id"] not in existing_map:
                result.append(item)
        return result

    # upsert (기본)
    result = list(existing)
    for item in incoming:
        pid = item["id"]
        if pid in existing_map:
            result[existing_map[pid]] = item
        else:
            result.append(item)
    return result


# ── 진입점 ─────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    filepath = Path(args[0])
    if not filepath.exists():
        print(f"ERROR: 파일 없음: {filepath}", file=sys.stderr)
        sys.exit(1)

    merge_strategy = "upsert"
    out_path = None

    i = 1
    while i < len(args):
        if args[i] == "--merge" and i + 1 < len(args):
            merge_strategy = args[i + 1]
            i += 2
        elif args[i] == "--out" and i + 1 < len(args):
            out_path = Path(args[i + 1])
            i += 2
        else:
            i += 1

    if merge_strategy not in ("upsert", "replace", "append"):
        print("ERROR: --merge는 upsert|replace|append 중 하나여야 합니다.", file=sys.stderr)
        sys.exit(1)

    if out_path is None:
        project_root = find_project_root(filepath.parent)
        out_path = project_root / "01-requirements" / "requirements.json"

    out_path.parent.mkdir(parents=True, exist_ok=True)

    version = parse_version_from_filename(filepath)

    # 기존 JSON 로드
    existing_data = {
        "_meta": {"version": None, "history": []},
        "_columns": {
            "functional": {
                "parent": list(FUNC_PARENT_FIELDS.values()),
                "sub": list(FUNC_SUB_FIELDS.values()),
            },
            "nonfunctional": {
                "parent": list(NONFUNC_PARENT_FIELDS.values()),
                "sub": list(NONFUNC_SUB_FIELDS.values()),
            },
        },
        "functional": [],
        "nonfunctional": [],
    }
    if out_path.exists():
        try:
            with open(out_path, encoding="utf-8") as f:
                loaded = json.load(f)
            existing_data["functional"]   = loaded.get("functional", [])
            existing_data["nonfunctional"] = loaded.get("nonfunctional", [])
            existing_data["_meta"]         = loaded.get("_meta", existing_data["_meta"])
        except Exception as e:
            print(f"WARNING: 기존 JSON 로드 실패, 새로 생성합니다: {e}", file=sys.stderr)

    # Excel 파싱
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"ERROR: Excel 파일 열기 실패: {e}", file=sys.stderr)
        sys.exit(1)

    any_imported = False
    for sheet_name in wb.sheetnames:
        sheet_type = detect_sheet_type(sheet_name)
        if sheet_type is None:
            print(f"INFO: 시트 '{sheet_name}' 건너뜀 (기능/비기능 판별 불가)")
            continue

        ws = wb[sheet_name]
        if sheet_type == "functional":
            incoming = parse_requirements_sheet(ws, FUNC_PARENT_FIELDS, FUNC_SUB_FIELDS)
        else:
            incoming = parse_requirements_sheet(ws, NONFUNC_PARENT_FIELDS, NONFUNC_SUB_FIELDS)

        if not incoming:
            print(f"INFO: 시트 '{sheet_name}'에 데이터가 없습니다.")
            continue

        merged = merge_requirements(existing_data[sheet_type], incoming, merge_strategy)
        existing_data[sheet_type] = merged
        any_imported = True
        print(f"✅ [{sheet_name}] {len(incoming)}개 파싱 → {merge_strategy} 적용")

    if not any_imported:
        print("ERROR: 처리할 데이터가 없습니다.", file=sys.stderr)
        sys.exit(1)

    # 메타 업데이트
    if version:
        existing_data["_meta"]["version"] = version
    existing_data["_meta"].setdefault("history", []).append({
        "version": version,
        "imported_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "source": filepath.name,
        "merge": merge_strategy,
    })

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(existing_data, f, ensure_ascii=False, indent=2)

    total_func    = len(existing_data.get("functional", []))
    total_nonfunc = len(existing_data.get("nonfunctional", []))
    print(f"\n📄 저장: {out_path}")
    print(f"   기능 요구사항:   {total_func}개")
    print(f"   비기능 요구사항: {total_nonfunc}개")


if __name__ == "__main__":
    main()
