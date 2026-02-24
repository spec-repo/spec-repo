#!/usr/bin/env python3
"""
요구사항정의서 Excel 파서.
xlsx 파일을 읽어 에이전트가 이해할 수 있는 구조화된 텍스트로 출력한다.

Usage:
    uv run --with openpyxl python3 parse-excel.py <excel-file>
    uv run --with openpyxl python3 parse-excel.py <excel-file> --sheet 기능_요구사항
    uv run --with openpyxl python3 parse-excel.py <excel-file> --summary
"""

import sys
import json
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.", file=sys.stderr)
    print("Run: uv run --with openpyxl python3 parse-excel.py <file>", file=sys.stderr)
    sys.exit(1)


def resolve_merged_cells(ws):
    """병합된 셀의 값을 모든 병합 범위에 채운다."""
    merge_values = {}
    for merge_range in ws.merged_cells.ranges:
        top_left = ws.cell(row=merge_range.min_row, column=merge_range.min_col)
        value = top_left.value
        for row in range(merge_range.min_row, merge_range.max_row + 1):
            for col in range(merge_range.min_col, merge_range.max_col + 1):
                merge_values[(row, col)] = value
    return merge_values


def get_cell_value(ws, row, col, merge_map):
    coord = (row, col)
    if coord in merge_map:
        return merge_map[coord]
    cell = ws.cell(row=row, column=col)
    return cell.value


def detect_header_rows(ws, merge_map, max_col):
    """1행과 2행이 헤더인지 판단하고 컬럼명 목록을 반환한다."""
    row1 = [get_cell_value(ws, 1, c, merge_map) for c in range(1, max_col + 1)]
    row2 = [get_cell_value(ws, 2, c, merge_map) for c in range(1, max_col + 1)]

    # 2행이 모두 row1과 같으면 단일 헤더 행
    has_subheader = any(
        row2[i] and row2[i] != row1[i] for i in range(len(row1))
    )

    headers = []
    if has_subheader:
        for i in range(max_col):
            r1 = str(row1[i]).strip() if row1[i] else ""
            r2 = str(row2[i]).strip() if row2[i] else ""
            if r2 and r2 != r1:
                headers.append(f"{r1} > {r2}" if r1 else r2)
            else:
                headers.append(r1)
        data_start_row = 3
    else:
        headers = [str(v).strip() if v else f"col_{i+1}" for i, v in enumerate(row1)]
        data_start_row = 2

    return headers, data_start_row


def parse_sheet(ws):
    """시트를 파싱해 헤더 + 행 데이터를 반환한다."""
    max_row = ws.max_row
    max_col = ws.max_column

    if max_row == 0 or max_col == 0:
        return {"headers": [], "rows": [], "total": 0}

    merge_map = resolve_merged_cells(ws)
    headers, data_start = detect_header_rows(ws, merge_map, max_col)

    rows = []
    for r in range(data_start, max_row + 1):
        row_vals = [get_cell_value(ws, r, c, merge_map) for c in range(1, max_col + 1)]

        # 빈 행 스킵
        if all(v is None or str(v).strip() == "" for v in row_vals):
            continue

        record = {}
        for i, h in enumerate(headers):
            val = row_vals[i] if i < len(row_vals) else None
            if val is not None:
                record[h] = str(val).strip()
        rows.append(record)

    return {"headers": headers, "rows": rows, "total": len(rows)}


def print_summary(data):
    """에이전트용 요약 텍스트 출력."""
    print(f"파일: {data['file']}")
    print(f"시트 수: {len(data['sheets'])}\n")

    for sheet_name, sheet_data in data['sheets'].items():
        total = sheet_data['total']
        headers = sheet_data['headers']
        rows = sheet_data['rows']

        print(f"━━━ [{sheet_name}] ━━━")
        print(f"컬럼 수: {len(headers)}, 데이터 행 수: {total}")
        print(f"컬럼 목록: {', '.join(headers)}")

        if rows:
            print(f"\n첫 번째 행 (샘플):")
            for k, v in list(rows[0].items())[:8]:
                print(f"  {k}: {v}")
            if len(rows[0]) > 8:
                print(f"  ... 외 {len(rows[0]) - 8}개 컬럼")

        # 요구사항 ID 목록
        id_cols = [h for h in headers if "ID" in h.upper() and "세부" not in h and ">" not in h]
        if id_cols and rows:
            id_col = id_cols[0]
            ids = [r[id_col] for r in rows if id_col in r]
            unique_ids = list(dict.fromkeys(ids))  # 순서 유지 dedup
            print(f"\n요구사항 ID 목록 ({len(unique_ids)}개): {', '.join(unique_ids[:20])}")
            if len(unique_ids) > 20:
                print(f"  ... 외 {len(unique_ids) - 20}개")
        print()


def main():
    args = sys.argv[1:]

    if not args or args[0] in ("-h", "--help"):
        print(__doc__)
        sys.exit(0)

    filepath = Path(args[0])
    if not filepath.exists():
        print(f"ERROR: 파일 없음: {filepath}", file=sys.stderr)
        sys.exit(1)

    sheet_filter = None
    if "--sheet" in args:
        idx = args.index("--sheet")
        if idx + 1 < len(args):
            sheet_filter = args[idx + 1]

    summary_mode = "--summary" in args

    try:
        wb = openpyxl.load_workbook(filepath, data_only=True)
    except Exception as e:
        print(f"ERROR: Excel 파일 열기 실패: {e}", file=sys.stderr)
        sys.exit(1)

    result = {"file": str(filepath), "sheets": {}}

    for name in wb.sheetnames:
        if sheet_filter and name != sheet_filter:
            continue
        result["sheets"][name] = parse_sheet(wb[name])

    if summary_mode:
        print_summary(result)
    else:
        print(json.dumps(result, ensure_ascii=False, indent=2, default=str))


if __name__ == "__main__":
    main()
